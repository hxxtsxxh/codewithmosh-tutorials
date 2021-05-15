# 'openpyxl.chart' - responsible for getting chart modules
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


# this function takes in a filename('xlsx')
# changes the prices initially found and writes corrected prices in another column
def process_workbook(filename):
    wb = xl.load_workbook(filename)  # loads the xml
    sheet = wb['Sheet1']  # accesses the sheet

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)


process_workbook('testXML.xlsx')
